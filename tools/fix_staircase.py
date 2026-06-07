# -*- coding: utf-8 -*-
"""Fix enemy difficulty staircase: boost 5 presets with incremental difficulty."""
import openpyxl

EXCEL_UNITS = "docs/excel/怪物配置_敌方单位明细表.xlsx"
EXCEL_PRESETS = "docs/excel/怪物配置_敌人预设表.xlsx"

# New values: (enemy_preset_id, slot_id, unit_id, count, star, role, notes)
# Counts boosted by ratio, stars boosted where gap is large
UPDATES = {
    "shadow_raiders": [
        ("enemy_1", "assassin", 17, 3, "前排爆发", "刺客暴击 检验后排承伤"),
        ("enemy_2", "wanderer", 14, 3, "副输出", "流浪者 中距输出"),
        ("enemy_3", "blacksmith", 17, 3, "辅助输出", "铁匠 ★3填充"),
        ("enemy_4", "light_mentor", 7, 4, "后排召唤", "光明导师 召唤幻影"),
    ],
    "garrison_line": [
        ("enemy_1", "garrison_guard", 7, 5, "前排核心", "卫戍协兵 全军增益联动×7"),
        ("enemy_2", "priest", 12, 4, "后排治疗", "牧师 ★4护盾支持"),
        ("enemy_3", "light_envoy", 7, 5, "后排支援", "莱特使者 信仰协同×7"),
        ("enemy_4", "wanderer", 15, 4, "副输出", "流浪者 ★4中距输出"),
    ],
    "abyss_vanguard": [
        ("enemy_1", "martial_master", 10, 5, "前排AOE", "武学大师 ★5范围伤害"),
        ("enemy_2", "light_mentor", 9, 5, "后排召唤", "光明导师 ★5召唤幻影"),
        ("enemy_3", "assassin", 18, 4, "刺客爆发", "刺客 ★4潜行暴击"),
        ("enemy_4", "garrison_guard", 7, 6, "辅助增益", "卫戍协兵 ★6全军buff"),
    ],
    "fallen_sanctum": [
        ("enemy_1", "light_envoy", 7, 6, "核心支援", "莱特使者 ★6信仰协同核心"),
        ("enemy_2", "garrison_guard", 8, 6, "前排承伤", "卫戍协兵 ★6高防联动"),
        ("enemy_3", "light_mentor", 10, 5, "后排召唤", "光明导师 ★5多单位召唤"),
        ("enemy_4", "royal_swordsman", 5, 6, "AOE输出", "皇家剑士 ★6战士AOE×5"),
    ],
    "doom_herald": [
        ("enemy_1", "royal_swordsman", 8, 6, "前排核心", "皇家剑士 ★6战士增益×8"),
        ("enemy_2", "echo_of_light", 4, 6, "后排输出", "莱特回响 开战全军增益×4"),
        ("enemy_3", "martial_master", 11, 5, "前排AOE", "武学大师 ★5数量压制"),
        ("enemy_4", "garrison_guard", 8, 6, "辅助增益", "卫戍协兵 ★6终局协同"),
    ],
}

wb = openpyxl.load_workbook(EXCEL_UNITS)
ws = wb.active

# Collect rows to update
preset_updates = {}
for row in range(2, ws.max_row + 1):
    pid = str(ws.cell(row=row, column=1).value or "").strip()
    if pid in UPDATES:
        if pid not in preset_updates:
            preset_updates[pid] = []
        preset_updates[pid].append(row)

for pid, new_slots in UPDATES.items():
    rows = preset_updates.get(pid, [])
    if len(rows) != len(new_slots):
        print(f"⚠ {pid}: expected {len(new_slots)} rows, found {len(rows)}")
        continue

    for i, row_num in enumerate(sorted(rows)):
        slot_id, unit_id, count, star, role, notes = new_slots[i]
        ws.cell(row=row_num, column=2).value = slot_id
        ws.cell(row=row_num, column=3).value = unit_id
        ws.cell(row=row_num, column=4).value = count
        ws.cell(row=row_num, column=5).value = star
        ws.cell(row=row_num, column=6).value = role
        ws.cell(row=row_num, column=7).value = notes
    print(f"✅ {pid} ({len(rows)} slots updated)")

wb.save(EXCEL_UNITS)
print(f"Saved {EXCEL_UNITS}")

# Update preset notes
NOTES_MAP = {
    "shadow_raiders": "Layer 4-5 爆发伤害检验 ★3-4强化版",
    "garrison_line": "Layer 5-7 协同压力 ★4-5群体增益强化",
    "abyss_vanguard": "Layer 7-9 AOE+召唤 ★5-6考验强化",
    "fallen_sanctum": "Layer 9-12 高协同 ★5-6生存检验强化",
    "doom_herald": "Layer 11-14 终局强度 ★5-6最终检验强化",
}

wb2 = openpyxl.load_workbook(EXCEL_PRESETS)
ws2 = wb2.active
for row in range(2, ws2.max_row + 1):
    pid = str(ws2.cell(row=row, column=1).value or "").strip()
    if pid in NOTES_MAP:
        ws2.cell(row=row, column=8).value = NOTES_MAP[pid]
        print(f"✅ Updated notes for {pid}")
wb2.save(EXCEL_PRESETS)
print(f"Saved {EXCEL_PRESETS}")
print("\nDone! Run sync + generate + verify.")