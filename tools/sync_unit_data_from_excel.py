# -*- coding: utf-8 -*-
"""Synchronize runtime unit data from docs/excel/unit_data.xlsx.

The workbook owns player-facing values and skill wording.  Structured skill
definitions live here explicitly so design changes remain reproducible instead
of being hand-edited in the generated JSON.
"""
from __future__ import annotations

import json
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "docs" / "excel" / "unit_data.xlsx"
OUTPUT = ROOT / "Assets" / "Resources" / "Data" / "unit_data.json"


def skill(kind: str, **values):
    return {"kind": kind, **values}


SKILL_OVERRIDES = {
    "small_merchant": (
        [skill("on_sell_every_n_gain_gold", threshold=2, value=1)],
        [skill("on_sell_every_n_gain_gold", threshold=2, value=2)], [], []),
    "elf": (
        [skill("while_on_board_on_entry_race_self_gain_attack", race="甘地", value=1)],
        [skill("while_on_board_on_entry_race_self_gain_attack", race="甘地", value=2)], [], []),
    "blacksmith": ([], [],
        [skill("battle_round_self_hp_loss_team_temp_attack", selfHpLoss=25, attack=1, intervalRounds=1)],
        [skill("battle_round_self_hp_loss_team_temp_attack", selfHpLoss=25, attack=2, intervalRounds=1)]),
    "knight": (
        [skill("round_end_if_race_count_self_gain_round_count", race="甘地", threshold=5, roundOffset=3)],
        [skill("round_end_if_race_count_self_gain_round_count", race="甘地", threshold=5, roundOffset=6)],
        [skill("battle_start_pounce_nearest_damage", attackMultiplier=1, forceCrit=True, stunTurns=1)],
        [skill("battle_start_pounce_nearest_damage", attackMultiplier=1, forceCrit=True, stunTurns=1, times=2)]),
    "priest": ([], [],
        [skill("battle_start_front_occupied_rows_shield", count=1, layers=1)],
        [skill("battle_start_front_occupied_rows_shield", count=2, layers=1)]),
    "wanderer": (
        [skill("while_on_board_every_n_entry_race_add_random_unit_to_hand", race="甘地", threshold=8, count=1)],
        [skill("while_on_board_every_n_entry_race_add_random_unit_to_hand", race="甘地", threshold=6, count=1)], [], []),
    "echo_of_light": (
        [skill("round_end_self_gain_attack_per_faith_count", faith="莱特", value=1)],
        [skill("round_end_self_gain_attack_per_faith_count", faith="莱特", value=2)],
        [skill("battle_start_team_count_per_faith_count", faith="莱特", value=10)],
        [skill("battle_start_team_count_per_faith_count", faith="莱特", value=15)]),
    "fire_elemental": (
        [skill("on_gain_power_self_gain_attack", value=4)],
        [skill("on_gain_power_self_gain_attack", value=8)],
        [skill("on_attack_multi_nearest_targets", targets=2, mode="same_damage")],
        [skill("on_attack_multi_nearest_targets", targets=3, mode="same_damage")]),
    "wind_elemental": (
        [skill("while_on_board_race_threshold_team_initiative", race="甘德", threshold=4, value=1)],
        [skill("while_on_board_race_threshold_team_initiative", race="甘德", threshold=4, value=2)],
        [skill("battle_start_tagged_units_consecutive_attacks", targetTag="元素", repeat=2)],
        [skill("battle_start_tagged_units_consecutive_attacks", targetTag="元素", repeat=3)]),
    "demon_lord": (
        [skill("on_gain_count_transfer_to_random_other_allies", count=3)],
        [skill("on_gain_count_transfer_to_random_other_allies", count=5)],
        [skill("battle_action_self_shield_if_none", layers=1)],
        [skill("battle_action_self_shield_if_none", layers=2)]),
    "elder_spirit": (
        [skill("leave_board_tagged_units_gain_stats", targetTags=["元素", "法师"], value=3)],
        [skill("leave_board_tagged_units_gain_stats", targetTags=["元素", "法师"], value=6)], [], []),
    "magic_dragon": ([], [],
        [skill("battle_start_summon_units", count=2, targetUnitId="fire_elemental", mode="highest_unit_count", threshold=10, ratio=1.0, summonUnitId="fire_elemental", temporary=True)],
        [skill("battle_start_summon_units", count=3, targetUnitId="fire_elemental", mode="highest_unit_count", threshold=10, ratio=1.0, summonUnitId="fire_elemental", temporary=True)]),
    "elite_ranger_rider": (
        [skill("round_end_self_gain_attack", value=1)],
        [skill("round_end_self_gain_attack", value=2)],
        [skill("on_attack_teleport_and_crit_chance", chance=0.5)],
        [skill("on_attack_teleport_and_crit_chance", chance=1.0)]),
    "sword_ranger": (
        [skill("on_receive_gift_self_gain_attack", value=4)],
        [skill("on_receive_gift_self_gain_attack", value=8)],
        [skill("on_damaged_survive_next_round_forest_gem", value=1)],
        [skill("on_damaged_survive_next_round_forest_gem", value=2)]),
    "river_captain": (
        [skill("on_entry_side_units_gift_forest_gem", value=1)],
        [skill("on_entry_side_units_gift_forest_gem", value=2)], [], []),
    "sky_watcher": (
        [skill("on_entry_any_unit_gain_forest_gem_self", value=1)],
        [skill("on_entry_any_unit_gain_forest_gem_self", value=2)],
        [skill("on_attack_chance_force_crit", chance=0.25, mode="absolute")],
        [skill("on_attack_chance_force_crit", chance=0.35, mode="absolute")]),
    "burrow_mole": (
        [skill("on_any_unit_gain_count_self_gift_and_self_multiple_gain_forest_gem", gift=1, threshold=10, gain=1)],
        [skill("on_any_unit_gain_count_self_gift_and_self_multiple_gain_forest_gem", gift=2, threshold=10, gain=1)],
        [skill("battle_start_stealth_first_attack_multiplier", attackMultiplier=2)],
        [skill("battle_start_stealth_first_attack_multiplier", attackMultiplier=5)]),
    "feather_guard": (
        [skill("round_end_if_no_forest_gem_in_hand_self_gain_attack", value=2)],
        [skill("round_end_if_no_forest_gem_in_hand_self_gain_attack", value=4)],
        [skill("on_attack_multi_nearest_targets", targets=2), skill("on_attack_mark_target_next_round_forest_gem_on_death", value=1)],
        [skill("on_attack_multi_nearest_targets", targets=3), skill("on_attack_mark_target_next_round_forest_gem_on_death", value=1)]),
    "cheetah": (
        [skill("on_gift_action_self_gain_count_every_n", threshold=5, value=4)],
        [skill("on_gift_action_self_gain_count_every_n", threshold=5, value=8)],
        [skill("on_death_next_round_self_count", value=5)],
        [skill("on_death_next_round_self_count", value=10)]),
    "twin_tower_mage": (
        [skill("round_end_gain_forest_gem_self", value=2)],
        [skill("round_end_gain_forest_gem_self", value=4)],
        [skill("battle_start_attach_to_highest_count_allies", count=1)],
        [skill("battle_start_attach_to_highest_count_allies", count=2)]),
    "xilinding": (
        [skill("on_gain_forest_gem_auto_gift_self_team_attack", value=4)],
        [skill("on_gain_forest_gem_auto_gift_self_team_attack", value=8)], [], []),
    "mire_fiend": (
        [skill("round_end_self_gift_forest_gem", value=3), skill("on_receive_gift_self_evolve", threshold=30, targetUnitId="blood_mire_fiend")],
        [skill("round_end_self_gift_forest_gem", value=4), skill("on_receive_gift_self_evolve", threshold=30, targetUnitId="blood_mire_fiend")], [], []),
    "laborer": (
        [skill("round_end_self_gain_attack", value=2), skill("on_sell_price_if_count_threshold", threshold=30, price=2)],
        [skill("round_end_self_gain_attack", value=3), skill("on_sell_price_if_count_threshold", threshold=30, price=4)], [], []),
    "snow_lion": (
        [skill("on_entry_board_tagged_units_gain_attack", targetTag="野兽", value=3)],
        [skill("on_entry_board_tagged_units_gain_attack", targetTag="野兽", value=6)],
        [skill("battle_start_pounce_nearest_damage", attackMultiplier=3)],
        [skill("battle_start_pounce_nearest_damage", attackMultiplier=6)]),
    "exorcist_mount": (
        [skill("round_end_devour_shop_extreme_count", targetMode="lowest", multiplier=1)],
        [skill("round_end_devour_shop_extreme_count", targetMode="highest", multiplier=1)],
        [skill("on_attack_count_fire_rain_area_dot", count=1, radius=2, duration=0, tick=0)],
        [skill("on_attack_count_fire_rain_area_dot", count=1, radius=3, duration=0, tick=0)]),
    "pain_flame": (
        [skill("on_leave_add_random_unit_from_pool", unitIds=["murloc_servant", "fire_elemental", "elf"], count=1)],
        [skill("on_leave_add_random_unit_from_pool", unitIds=["murloc_servant", "fire_elemental", "elf", "caller"], count=1)], [], []),
    "shadow_butcher": (
        [skill("while_on_board_on_entry_race_devour_shop_gain_attack", race="甘格尔", threshold=2, multiplier=1)],
        [skill("while_on_board_on_entry_race_devour_shop_gain_attack", race="甘格尔", threshold=2, multiplier=2)],
        [skill("on_attack_chance_self_shield_no_stack", chance=0.15, layers=1)],
        [skill("on_attack_chance_self_shield_no_stack", chance=0.25, layers=1)]),
    "ger_giant_beast": (
        [skill("on_devour_self_gain_attack", value=4)],
        [skill("on_devour_self_gain_attack", value=8)],
        [skill("first_attack_devour_until_self_death_or_no_enemies")],
        [skill("first_attack_devour_until_self_death_or_no_enemies")]),
    "novice_elementalist": (
        [skill("while_on_board_on_entry_same_id_tagged_units_gain_power", tag="元素", value=1)],
        [skill("while_on_board_on_entry_same_id_tagged_units_gain_power", tag="元素", value=2)],
        [skill("battle_start_summon_units", count=1, summonUnitId="fire_elemental", targetUnitId="fire_elemental", mode="highest_unit_count", threshold=10, ratio=1.0, temporary=True)],
        [skill("battle_start_summon_units", count=2, summonUnitId="fire_elemental", targetUnitId="fire_elemental", mode="highest_unit_count", threshold=10, ratio=1.0, temporary=True)]),
    "water_elemental": (
        [skill("while_on_board_per_ally_id_buff_type_attack", allyId="fire_elemental", targetTag="元素", attack=2)],
        [skill("while_on_board_per_ally_id_buff_type_attack", allyId="fire_elemental", targetTag="元素", attack=4)], [], []),
    "element_master": ([], [],
        [skill("on_attack_summon_locked_units", summonUnitId="fire_elemental", count=1, value=10, temporary=True)],
        [skill("on_attack_summon_locked_units", summonUnitId="fire_elemental", count=2, value=10, temporary=True)]),
    "ger_officer": (
        [skill("on_entry_random_board_tag_gain_own_initial_count", targetTag="野兽", count=1)],
        [skill("on_entry_random_board_tag_gain_own_initial_count", targetTag="野兽", count=2)], [], []),
}


def main():
    workbook = openpyxl.load_workbook(WORKBOOK, data_only=True, read_only=True)
    sheet = workbook["单位数值设定"]
    with OUTPUT.open("r", encoding="utf-8-sig") as handle:
        units = json.load(handle)
    by_name = {unit["name"]: unit for unit in units}

    field_columns = {
        "star": 2, "race": 3, "typeLabel": 4, "faith": 5,
        "startCount": 7, "hpPerUnit": 8, "attack": 9, "defense": 10,
        "damageMin": 11, "damageMax": 12, "initiative": 13, "speed": 14,
        "morale": 15, "luck": 16, "attackRange": 17, "size": 18,
        "firstPurchaseHp": 19, "firstPurchaseAverageDamage": 20,
        "limit": 21,
    }
    text_columns = {
        "talentText": 22, "goldTalentText": 23,
        "battleText": 24, "goldBattleText": 25,
    }

    seen = set()
    for row in range(3, sheet.max_row + 1):
        name = str(sheet.cell(row, 1).value or "").strip()
        if not name:
            continue
        if name not in by_name:
            raise KeyError(f"Workbook unit is missing from JSON: {name}")
        unit = by_name[name]
        seen.add(name)

        for key, column in field_columns.items():
            value = sheet.cell(row, column).value
            if key == "typeLabel" and str(value or "").strip() == "-":
                value = ""
            if key == "limit":
                value = 0 if str(value or "").strip() == "-" else int(value)
            if key == "attackRange":
                value = float(value)
            unit[key] = value
        unit["hp"] = unit["hpPerUnit"]

        tag_value = sheet.cell(row, 6).value
        if tag_value is None or str(tag_value).strip() == "-":
            unit["tags"] = []
        else:
            unit["tags"] = [part.strip() for part in str(tag_value).replace("，", ",").split(",") if part.strip()]

        for key, column in text_columns.items():
            value = sheet.cell(row, column).value
            unit[key] = "—" if value is None or not str(value).strip() else str(value).strip()
        unit["skillText"] = "\n".join(unit[key] for key in text_columns)

    missing = sorted(set(by_name) - seen)
    if missing:
        raise ValueError(f"JSON units missing from workbook: {missing}")

    for unit_id, definitions in SKILL_OVERRIDES.items():
        unit = next((item for item in units if item.get("id") == unit_id), None)
        if unit is None:
            raise KeyError(f"Skill override references unknown unit id: {unit_id}")
        unit["talents"], unit["goldTalents"], unit["battleSkills"], unit["goldBattleSkills"] = definitions

    # The table has no battle/talent effect for these definitions.
    for unit_id in ("ger_beast",):
        unit = next(item for item in units if item.get("id") == unit_id)
        unit["battleSkills"] = []
        unit["goldBattleSkills"] = []

    # Remove hidden Ger-beast attack synchronization and use death-only summoning.
    tamer = next(item for item in units if item.get("id") == "beast_tamer")
    tamer["battleSkills"] = [skill("battle_start_summon_units", summonUnitId="ger_beast", count=1, temporary=True)]
    tamer["goldBattleSkills"] = [skill("battle_start_summon_units", summonUnitId="ger_beast", count=2, temporary=True)]
    rider = next(item for item in units if item.get("id") == "beast_rider")
    rider["battleSkills"] = [skill("on_death_summon_units", summonUnitId="ger_beast", count=1, mode="origin_then_nearest", temporary=True)]
    rider["goldBattleSkills"] = [skill("on_death_summon_units", summonUnitId="ger_beast", count=2, mode="origin_then_nearest", temporary=True)]

    beastmaster = next(item for item in units if item.get("id") == "witch_beastmaster")
    beastmaster["goldBattleSkills"] = [skill("on_attack_multi_nearest_targets", targets=5)]

    mentor = next(item for item in units if item.get("id") == "light_mentor")
    for definitions in (mentor["battleSkills"], mentor["goldBattleSkills"]):
        for definition in definitions:
            definition["value"] = 1
            definition["disableAttack"] = True
            definition["disableMovement"] = True
            definition["temporary"] = True

    earth = next(item for item in units if item.get("id") == "earth_elemental")
    earth["talents"] = [skill("on_leave_tag_count_tagged_units_gain_power", tag="魔灵", targetTag="元素", threshold=4, value=4)]
    earth["goldTalents"] = [skill("on_leave_tag_count_tagged_units_gain_power", tag="魔灵", targetTag="元素", threshold=2, value=4)]

    with OUTPUT.open("w", encoding="utf-8", newline="\n") as handle:
        json.dump(units, handle, ensure_ascii=False, indent=4)
        handle.write("\n")

    print(f"Synchronized {len(units)} units from {WORKBOOK.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
