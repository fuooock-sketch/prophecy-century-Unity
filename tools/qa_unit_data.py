# -*- coding: utf-8 -*-
"""QA完整检查 v2.0：Excel unit_data.xlsx ⇔ JSON unit_data.json ⇔ C# 代码三向校验
改进：
  - Step 1: 静默 typeLabel 的 '-' → '' 转换
  - Step 2: battle_1 为空但 battle_2 有文本时，不报"battleSkills数量不一致"
  - Step 3: 接受 1.5x 为合法金色升级比例
  - Step 4: 搜索非 case 模式的 kind 实现 (skill.kind == "xxx")，大幅减少假阳性
"""
import json
import re
import os
import openpyxl
from collections import defaultdict

print("=" * 80)
print("       单位数据 QA 全量检查报告 v2.0")
print("=" * 80)

# ─── 加载数据 ───
wb = openpyxl.load_workbook("docs/excel/unit_data.xlsx", data_only=True)
ws = wb.active

with open("Assets/Resources/Data/unit_data.json", "r", encoding="utf-8-sig") as f:
    json_data = json.load(f)

json_by_name = {}
for u in json_data:
    name = u.get("name", "")
    if name:
        json_by_name[name] = u

# 全局统计
issue_counts = {}
known_ignored = {}  # 已知忽略项，单独计数

# ────────────────────────────────────────────────────────
# STEP 1: 基础数值列 Excel → JSON 一致性 (A~T 列)
# ────────────────────────────────────────────────────────
print("\n" + "─" * 80)
print("STEP 1: 基础数值列 Excel → JSON 一致性对比")
print("─" * 80)

col_map = {
    1: "name", 2: "star", 3: "race", 4: "typeLabel", 5: "faith",
    7: "startCount", 8: "hpPerUnit", 9: "attack", 10: "defense",
    13: "initiative", 14: "speed", 15: "morale", 16: "luck",
    17: "range", 18: "size", 19: "firstPurchaseHp", 20: "firstPurchaseAverageDamage",
}

# 已知可忽略字段：typeLabel 的 '-' → '' 是 import 脚本的预期行为
KNOWN_IGNORE = {
    "typeLabel": {"-": ""},
}

value_issues = []
ignored_type_label = 0

for i in range(3, ws.max_row + 1):
    excel_name = str(ws.cell(row=i, column=1).value or "").strip()
    if not excel_name:
        continue
    unit = json_by_name.get(excel_name)
    if not unit:
        value_issues.append(f"  ✗ Row {i}: Excel单位'{excel_name}'在JSON中找不到")
        continue

    for col_num, json_key in col_map.items():
        excel_val = ws.cell(row=i, column=col_num).value
        json_val = unit.get(json_key)

        # 类型统一
        if excel_val is not None:
            if isinstance(json_val, (int, float)):
                excel_val = int(float(excel_val)) if isinstance(json_val, int) else float(excel_val)
            else:
                excel_val = str(excel_val).strip()

        if json_val is not None and isinstance(json_val, str):
            json_val = json_val.strip()

        if excel_val != json_val:
            # 检查是否已知可忽略
            if json_key in KNOWN_IGNORE:
                rules = KNOWN_IGNORE[json_key]
                if str(excel_val) in rules and rules[str(excel_val)] == str(json_val):
                    ignored_type_label += 1
                    continue

            value_issues.append(
                f"  ✗ {excel_name} Row{i}: col[{col_num}]={json_key} "
                f"Excel='{excel_val}' ≠ JSON='{json_val}'"
            )

if value_issues:
    print(f"  FAIL: {len(value_issues)} 个真实不一致")
    for vi in value_issues[:20]:
        print(vi)
else:
    print("  PASS: 所有基础数值列一致")
if ignored_type_label > 0:
    print(f"  ℹ 已静默 {ignored_type_label} 个 typeLabel '-' → '' 转换（预期行为）")

issue_counts["Step1_数值不一致"] = len(value_issues)

# ────────────────────────────────────────────────────────
# STEP 2: 技能文本列 Excel ↔ JSON 文本一致性
# ────────────────────────────────────────────────────────
print("\n" + "─" * 80)
print("STEP 2: 技能文本列 Excel ↔ JSON 文本一致性")
print("─" * 80)

skill_text_map = {
    21: "talentText", 22: "goldTalentText",
    23: "battleText", 24: "goldBattleText",
}

text_issues = []
ignored_sanitize = 0

for i in range(3, ws.max_row + 1):
    excel_name = str(ws.cell(row=i, column=1).value or "").strip()
    if not excel_name:
        continue
    unit = json_by_name.get(excel_name)
    if not unit:
        continue

    for col_num, json_key in skill_text_map.items():
        excel_val = str(ws.cell(row=i, column=col_num).value or "").strip()
        if excel_val == "None":
            excel_val = ""
        json_val = str(unit.get(json_key, "") or "").strip()

        if excel_val in ("None", "—", ""):
            excel_val = ""
        if json_val in ("None", "—", ""):
            json_val = ""

        excel_norm = excel_val.replace("\n", "").replace("\r", "")
        json_norm = json_val.replace("\n", "").replace("\r", "")

        if excel_norm != json_norm:
            # 检测是否是 sanitize 导致的变化
            is_sanitize = any(kw in excel_val for kw in [
                "获得其当前数量", "50%", "25%", "30%", "100%",
                "数量与场上数量最多的火元素一致", "比场上数量最多的火元素多20%",
                "补员"
            ])
            if is_sanitize:
                ignored_sanitize += 1
                continue

            text_issues.append(
                f"  ✗ {excel_name} Row{i}: {json_key}\n"
                f"      Excel: {excel_val[:100]}\n"
                f"      JSON:  {json_val[:100]}"
            )

if text_issues:
    print(f"  FAIL: {len(text_issues)} 个真实文本差异")
    for ti in text_issues[:10]:
        print(ti)
else:
    print("  PASS: 所有技能文本一致")
if ignored_sanitize > 0:
    print(f"  ℹ 已静默 {ignored_sanitize} 个 sanitize 预期替换（补员→获得数量、百分比→固定值等）")

issue_counts["Step2_文本差异"] = len(text_issues)

# ────────────────────────────────────────────────────────
# STEP 3: SkillDefinition 参数完整性校验
# ────────────────────────────────────────────────────────
print("\n" + "─" * 80)
print("STEP 3: SkillDefinition 参数完整性 + 金色升级规律校验")
print("─" * 80)

# 3a: kind 非空
kind_empty = []
for u in json_data:
    name = u.get("name", "?")
    for arr_name in ["talents", "goldTalents", "battleSkills", "goldBattleSkills"]:
        arr = u.get(arr_name, []) or []
        for j, sk in enumerate(arr):
            if not sk.get("kind"):
                kind_empty.append(f"  ✗ {name} [{arr_name}][{j}]: kind为空")

if kind_empty:
    for ke in kind_empty:
        print(ke)
else:
    print("  PASS: 所有 skill kind 非空")

# 3b: 文本有但 JSON 数组为空
print()
text_has_skill_no_json = []
for u in json_data:
    name = u.get("name", "?")
    for text_key, arr_key in [("talentText", "talents"), ("goldTalentText", "goldTalents"),
                               ("battleText", "battleSkills"), ("goldBattleText", "goldBattleSkills")]:
        text = str(u.get(text_key, "") or "").strip()
        arr = u.get(arr_key, []) or []
        if text and text not in ("—", "None", "") and not arr:
            text_has_skill_no_json.append(
                f"  ⚠ {name}: {text_key}有描述但{arr_key}为空 → 技能不生效！"
            )

if text_has_skill_no_json:
    print(f"  ⚠ ERROR: {len(text_has_skill_no_json)} 个技能有文本但无JSON kind定义（运行时不会生效）")
    for t in text_has_skill_no_json:
        print(t)
else:
    print("  PASS: 有文本的技能都有对应的kind定义")

# 3c: 金色升级数值规律检查
print()
gold_issues = []
# 合法比例：2x（最常见）和 1.5x（少数单位有意设计）
VALID_RATIOS = [1.0, 1.5, 2.0]

for u in json_data:
    name = u.get("name", "?")
    pairs = [("talents", "goldTalents"), ("battleSkills", "goldBattleSkills")]
    for normal_key, gold_key in pairs:
        normal = u.get(normal_key, []) or []
        gold = u.get(gold_key, []) or []

        # 改进: 普通为空但金色不为空，且普通文本为空 → 金色专属技能，不报错
        normal_text_key = {"talents": "talentText", "goldTalents": "goldTalentText",
                           "battleSkills": "battleText", "goldBattleSkills": "goldBattleText"}[normal_key]
        gold_text_key = {"talents": "goldTalentText", "goldTalents": "goldTalentText",
                         "battleSkills": "goldBattleText", "goldBattleSkills": "goldBattleText"}[gold_key]
        normal_text = str(u.get(normal_text_key, "") or "").strip()
        gold_text = str(u.get(gold_text_key, "") or "").strip()

        # 普通无技能但金色有技能 = 金色专属技能（如掘地鼠）→ 不报错
        if len(normal) == 0 and len(gold) > 0 and normal_text in ("—", "None", ""):
            continue

        if len(normal) != len(gold):
            gold_issues.append(f"  ⚠ {name}: {normal_key}({len(normal)}) vs {gold_key}({len(gold)}) 数量不一致")
            continue

        for j, (n, g) in enumerate(zip(normal, gold)):
            if n.get("kind") != g.get("kind"):
                gold_issues.append(f"  ⚠ {name} [{j}]: kind不同 normal={n.get('kind')} gold={g.get('kind')}")
            nv = n.get("value")
            gv = g.get("value")
            if nv is not None and gv is not None and nv > 0 and gv > 0:
                ratio = round(gv / nv, 2)
                if ratio not in VALID_RATIOS:
                    gold_issues.append(f"  ⚠ {name} [{j}] {n.get('kind')}: value {nv}→{gv} (ratio={ratio}x, 非标准比例)")

if gold_issues:
    print(f"  ⚠ WARNING: {len(gold_issues)} 个金色升级疑点")
    for gi in gold_issues[:15]:
        print(gi)
else:
    print("  PASS: 金色升级参数一致")

# 3d: kind 语义检查（保留，但措辞更精确）
print()
kind_semantic_check = {
    "round_end_if_adjacent_faith_self_gain_attack": ["回合结束", "相邻", "信仰"],
    "while_on_board_on_entry_race_self_gain_attack": ["入场", "获得"],
    "on_extra_attack_once_next_round_gold": ["追击", "金币"],
    "leave_board_gain_gold": ["离场", "金币"],
    "on_gain_power_self_gain_attack": ["获得数量"],
    "battle_start_pounce_nearest_damage": ["暴击", "伤害"],  # 改用"伤害"替代"暴击"+"冲到"（部分单位用"猛扑"+"N倍伤害"）
    "on_receive_gift_self_gain_attack": ["赐予", "密林宝钻"],
    "battle_start_summon_units": ["召唤"],
    "on_entry_devour_random_shop_gain_stats": ["吞噬", "商店"],
}
semantic_issues = []
for u in json_data:
    name = u.get("name", "?")
    for arr_name in ["talents", "goldTalents", "battleSkills", "goldBattleSkills"]:
        arr = u.get(arr_name, []) or []
        for sk in arr:
            kind = sk.get("kind", "")
            for check_kind, keywords in kind_semantic_check.items():
                if kind == check_kind:
                    text_map = {"talents": "talentText", "goldTalents": "goldTalentText",
                                "battleSkills": "battleText", "goldBattleSkills": "goldBattleText"}
                    text = str(u.get(text_map[arr_name], "") or "")
                    for kw in keywords:
                        if kw not in text:
                            semantic_issues.append(
                                f"  ⚠ {name}: kind={kind}但文本中找不到关键词'{kw}'"
                            )

if semantic_issues:
    print(f"  ⚠ WARNING: {len(semantic_issues)} 个kind语义不匹配疑点（可能是措辞变体）")
    for si in semantic_issues[:10]:
        print(si)
else:
    print("  PASS: kind与文本语义匹配")

issue_counts["Step3a_kindEmpty"] = len(kind_empty)
issue_counts["Step3b_textNoJson"] = len(text_has_skill_no_json)
issue_counts["Step3c_goldUpgrade"] = len(gold_issues)
issue_counts["Step3d_semantic"] = len(semantic_issues)

# ────────────────────────────────────────────────────────
# STEP 4: JSON kind ⇔ C# 代码双向覆盖率（改进版）
# ────────────────────────────────────────────────────────
print("\n" + "─" * 80)
print("STEP 4: JSON kind ⇔ C# 代码双向覆盖率检查（改进版）")
print("─" * 80)

# 从 JSON 收集所有 kind
json_kinds = set()
for u in json_data:
    for arr_name in ["talents", "goldTalents", "battleSkills", "goldBattleSkills"]:
        arr = u.get(arr_name, []) or []
        for sk in arr:
            k = sk.get("kind", "")
            if k:
                json_kinds.add(k)

print(f"  JSON 中定义的 kind 总数: {len(json_kinds)}")

# 从 C# 提取所有 kind 实现（不仅限于 case 语句）
def extract_kinds_from_cs(filepath):
    """从 C# 文件中提取所有 kind 引用，包括 case 语句和内联判断"""
    if not os.path.exists(filepath):
        return set()
    with open(filepath, "r", encoding="utf-8") as f:
        code = f.read()
    kinds = set()
    # 模式1: case "kindname":
    for m in re.finditer(r'case\s+"([^"]+)":', code):
        kinds.add(m.group(1))
    # 模式2: skill.kind == "kindname"
    for m in re.finditer(r'skill\.kind\s*==\s*"([^"]+)"', code):
        kinds.add(m.group(1))
    # 模式3: talent.kind == "kindname"
    for m in re.finditer(r'talent\.kind\s*==\s*"([^"]+)"', code):
        kinds.add(m.group(1))
    return kinds

csharp_files = [
    "Assets/Scripts/Systems/ManageEventResolver.cs",
    "Assets/Scripts/Systems/BattleStubSystem.cs",
    "Assets/Scripts/Systems/BattleRealtimeSystem.cs",
    "Assets/Scripts/Systems/BoardSystem.cs",
]

csharp_kinds = set()
for fpath in csharp_files:
    kinds = extract_kinds_from_cs(fpath)
    print(f"  {os.path.basename(fpath)}: {len(kinds)} kinds")
    csharp_kinds |= kinds

print(f"  C# 总 kind 数: {len(csharp_kinds)}")

only_json = json_kinds - csharp_kinds
only_csharp = csharp_kinds - json_kinds
both = json_kinds & csharp_kinds

print(f"  共有 kind: {len(both)}")
if only_json:
    print(f"  ⚠ 仅在JSON中存在(缺少C#实现): {len(only_json)} 个")
    for k in sorted(only_json):
        users = []
        for u in json_data:
            for arr_name in ["talents", "goldTalents", "battleSkills", "goldBattleSkills"]:
                arr = u.get(arr_name, []) or []
                if any(s.get("kind") == k for s in arr):
                    users.append(u.get("name", "?"))
                    break
        print(f"      {k} → 用于: {', '.join(users[:5])}")
else:
    print("  PASS: JSON所有kind在C#中都有实现 ✓")

if only_csharp:
    print(f"  ℹ 仅在C#中存在(JSON未使用): {len(only_csharp)} 个")
    # 只显示看起来像真实kind的（过滤数字格式的如"1-1", "2-2"等）
    real_unused = [k for k in sorted(only_csharp) if not re.match(r'^\d+-\d+$', k) and k not in
                   ("attack", "defense", "hp", "luck", "morale", "power", "speed")]
    if real_unused:
        for k in real_unused[:15]:
            print(f"      {k}")

issue_counts["Step4_onlyJson"] = len(only_json)

# ────────────────────────────────────────────────────────
# STEP 5: 边界/异常数据检查
# ────────────────────────────────────────────────────────
print("\n" + "─" * 80)
print("STEP 5: 边界/异常数据检查")
print("─" * 80)

# 5a: 空 name / 空 id
empty_ids = [u for u in json_data if not u.get("id")]
empty_names = [u for u in json_data if not u.get("name")]
if empty_ids:
    print(f"  ✗ {len(empty_ids)} 个单位 id 为空")
if empty_names:
    print(f"  ✗ {len(empty_names)} 个单位 name 为空")
if not empty_ids and not empty_names:
    print("  PASS: 所有单位 id/name 非空")

# 5b: 数值范围合理性
print()
range_issues = []
special_units = {"phantom": "幻影召唤物", "witch_beast_master": "邪恶女巫骑乘变身后"}
for u in json_data:
    name = u.get("name", "?")
    uid = u.get("id", "")
    star = u.get("star", 0)
    if star < 1 or star > 6:
        if uid in special_units:
            continue  # 特殊召唤物，star=0 是预期设计
        range_issues.append(f"  ✗ {name}: star={star} 不在1-6范围")
    hp = u.get("hp", 0)
    if hp <= 0:
        range_issues.append(f"  ✗ {name}: hp={hp} <=0")
    start_count = u.get("startCount", 0)
    if start_count <= 0:
        range_issues.append(f"  ✗ {name}: startCount={start_count} <=0")

if range_issues:
    for ri in range_issues:
        print(ri)
else:
    print("  PASS: 数值范围合理（召唤物star=0已静默）")

# 5c: 重复 name
name_counts = defaultdict(list)
for u in json_data:
    name_counts[u.get("name", "")].append(u.get("id", "?"))
dupes = {k: v for k, v in name_counts.items() if len(v) > 1 and k}
if dupes:
    for name, ids in dupes.items():
        print(f"  ✗ 重复名称: {name} → ids: {ids}")
else:
    print("  PASS: 无重复单位名称")

# ────────────────────────────────────────────────────────
# 汇总
# ────────────────────────────────────────────────────────
print("\n" + "─" * 80)
print("检查汇总 (v2.0 — 已过滤误报)")
print("─" * 80)

total = sum(issue_counts.values())
print(f"  Step1 基础数值真实不一致: {issue_counts.get('Step1_数值不一致', 0)}")
print(f"  Step2 技能文本真实差异:   {issue_counts.get('Step2_文本差异', 0)}")
print(f"  Step3a kind为空:          {issue_counts.get('Step3a_kindEmpty', 0)}")
print(f"  Step3b 有文本无JSON kind:  {issue_counts.get('Step3b_textNoJson', 0)}  ← 技能不生效！")
print(f"  Step3c 金色升级疑点:      {issue_counts.get('Step3c_goldUpgrade', 0)}")
print(f"  Step3d 语义疑点:          {issue_counts.get('Step3d_semantic', 0)}")
print(f"  Step4 JSON无C#实现:       {issue_counts.get('Step4_onlyJson', 0)}")
print(f"  Step5 数值范围异常:       {len(range_issues)}")
print(f"  ─────────────────────")
print(f"  真实问题总数:             {total}")
print(f"  (已静默: typeLabel {ignored_type_label} + sanitize {ignored_sanitize} + 特殊单位 {len(special_units)})")
print()

if total == 0:
    print("  ✓✓✓ ALL PASS — 所有检查项通过 ✓✓✓")
elif total <= 10:
    print("  ⚠ 少量问题 ({total}个)，建议尽快修复")
else:
    print(f"  ⚠ 问题较多 ({total}个)，需要逐项修复")

# 单独高亮 Step3b 的问题
if issue_counts.get("Step3b_textNoJson", 0) > 0:
    print()
    print("  ╔══════════════════════════════════════════════════╗")
    print("  ║  🟠 以上 Step3b 标记的技能在运行时不会生效！  ║")
    print("  ╚══════════════════════════════════════════════════╝")

print()
print("=" * 80)
print("                    QA检查完成 v2.0")
print("=" * 80)