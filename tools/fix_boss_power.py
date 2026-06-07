# -*- coding: utf-8 -*-
"""Fix Boss abyss_lord power from 901 → target ~1500.
Replaces light_mentor with martial_master, boosts all units.
"""
import openpyxl

EXCEL_UNITS = "docs/excel/怪物配置_敌方单位明细表.xlsx"
EXCEL_PRESETS = "docs/excel/怪物配置_敌人预设表.xlsx"

# ===== 1. Update enemy unit details =====
wb = openpyxl.load_workbook(EXCEL_UNITS)
ws = wb.active

# Find abyss_lord rows and update them
new_units = [
    # (row_idx, slot_id, unit_id, count, star, role, notes)
    # Original rows are in order: enemy_1..enemy_4
    ("enemy_1", "echo_of_light", 6, 6, "Boss核心", "莱特回响 终局全军爆发增益 ×6"),
    ("enemy_2", "royal_swordsman", 10, 6, "副输出", "皇家剑士 战士标签增益 ×10"),
    ("enemy_3", "martial_master", 12, 5, "AOE压制", "武学大师 范围伤害压场"),
    ("enemy_4", "assassin", 18, 4, "暗刺", "刺客 潜行暴击 终局暗杀者 ★4强化"),
]

# Find rows for abyss_lord
abyss_rows = []
for row in range(2, ws.max_row + 1):
    if ws.cell(row=row, column=1).value == "abyss_lord":
        abyss_rows.append(row)

print(f"Found {len(abyss_rows)} abyss_lord rows at rows: {abyss_rows}")

if len(abyss_rows) != 4:
    print(f"WARNING: Expected 4 rows, found {len(abyss_rows)}. Please check Excel.")
    # Try to find them
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if val and "abyss" in str(val):
            print(f"  Row {row}: {val}")

# Update each row
for i, row_num in enumerate(abyss_rows):
    slot_id, unit_id, count, star, role, notes = new_units[i]
    ws.cell(row=row_num, column=2).value = slot_id
    ws.cell(row=row_num, column=3).value = unit_id
    ws.cell(row=row_num, column=4).value = count
    ws.cell(row=row_num, column=5).value = star
    ws.cell(row=row_num, column=6).value = role
    ws.cell(row=row_num, column=7).value = notes
    print(f"  Slot {slot_id}: {unit_id} ×{count} ★{star} → {role}")

wb.save(EXCEL_UNITS)
print(f"✅ Updated {EXCEL_UNITS}")

# ===== 2. Update enemy preset notes =====
wb2 = openpyxl.load_workbook(EXCEL_PRESETS)
ws2 = wb2.active

for row in range(2, ws2.max_row + 1):
    if ws2.cell(row=row, column=1).value == "abyss_lord":
        ws2.cell(row=row, column=8).value = "Boss全明星阵容 ★4-6强化版 目标战力~1500"
        print(f"✅ Updated abyss_lord preset notes at row {row}")
        break

wb2.save(EXCEL_PRESETS)
print(f"✅ Updated {EXCEL_PRESETS}")
print("\nDone! Run sync_csv_from_excel.py → generate_config_json.py to apply.")