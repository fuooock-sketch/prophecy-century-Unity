# -*- coding: utf-8 -*-
"""分析 unit_data.xlsx 中 U/V/W/X 四列技能设计底层架构"""
import openpyxl
from collections import defaultdict

wb = openpyxl.load_workbook("docs/excel/unit_data.xlsx", data_only=True)
ws = wb.active

# ---------- 1. 读取数据 ----------
units = []
for i in range(3, ws.max_row + 1):
    u = ws.cell(row=i, column=1).value
    lv = ws.cell(row=i, column=2).value
    race = ws.cell(row=i, column=3).value
    utype = ws.cell(row=i, column=4).value
    faith = ws.cell(row=i, column=5).value
    tag = ws.cell(row=i, column=6).value
    t1 = str(ws.cell(row=i, column=21).value) if ws.cell(row=i, column=21).value else ""
    t2 = str(ws.cell(row=i, column=22).value) if ws.cell(row=i, column=22).value else ""
    b1 = str(ws.cell(row=i, column=23).value) if ws.cell(row=i, column=23).value else ""
    b2 = str(ws.cell(row=i, column=24).value) if ws.cell(row=i, column=24).value else ""
    units.append(
        {
            "unit": u,
            "level": lv,
            "race": race,
            "type": utype,
            "faith": faith,
            "tag": tag,
            "talent_1": t1,
            "talent_2": t2,
            "battle_1": b1,
            "battle_2": b2,
        }
    )

# ---------- 统计 ----------
total = len(units)

# 每个单位的技能列数量（非 "—" / "None" / 空）
def has_skill(text):
    return text and text != "—" and text != "None"

talent1_count = sum(1 for u in units if has_skill(u["talent_1"]))
talent2_count = sum(1 for u in units if has_skill(u["talent_2"]))
battle1_count = sum(1 for u in units if has_skill(u["battle_1"]))
battle2_count = sum(1 for u in units if has_skill(u["battle_2"]))
both_talent = sum(1 for u in units if has_skill(u["talent_1"]) and has_skill(u["talent_2"]))
both_battle = sum(1 for u in units if has_skill(u["battle_1"]) and has_skill(u["battle_2"]))
only_talent = sum(1 for u in units if (has_skill(u["talent_1"]) or has_skill(u["talent_2"])) and not has_skill(u["battle_1"]) and not has_skill(u["battle_2"]))
only_battle = sum(1 for u in units if (has_skill(u["battle_1"]) or has_skill(u["battle_2"])) and not has_skill(u["talent_1"]) and not has_skill(u["talent_2"]))
both_phases = sum(1 for u in units if (has_skill(u["talent_1"]) or has_skill(u["talent_2"])) and (has_skill(u["battle_1"]) or has_skill(u["battle_2"])))

print("=" * 70)
print("            单位技能系统 底层架构分析报告")
print("=" * 70)
print(f"数据来源: docs/excel/unit_data.xlsx → Sheet '单位数值设定'")
print(f"分析列: U=talent_1, V=talent_2, W=battle_1, X=battle_2")
print(f"有效单位数: {total} (Row 3 ~ Row {2+total})")
print()

# ──────────────────────────────────────────────
# 一、双阶段架构
# ──────────────────────────────────────────────
print("━" * 70)
print("一、核心架构：双阶段技能分离（Two-Phase Skill System）")
print("━" * 70)
print()
print("  本系统的底层设计哲学是将技能严格划分为两个互不重叠的阶段：")
print()
print("  【阶段A — 运营阶段 Operations Phase】(talent_1 / talent_2)")
print("    触发时机：在准备回合/商店阶段生效，属于经济与成长层。")
print("    设计目标：影响资源获取（金币/手牌）、数量成长（获得数量）、")
print("              种族协同、进阶解锁、商店操控。")
print("    对应列：U列 talent_1（普通形态）, V列 talent_2（金色升级形态）")
print()
print("  【阶段B — 战斗阶段 Battle Phase】(battle_1 / battle_2)")
print("    触发时机：在双方部队进入战斗结算时生效，属于战术层。")
print("    设计目标：影响伤害输出、暴击、护盾、召唤、位移、控场、")
print("              AOE、反击、追击等战斗内行为。")
print("    对应列：W列 battle_1（普通形态）, X列 battle_2（金色升级形态）")
print()

print(f"  技能分布统计：")
print(f"    有运营技能 (talent) 的单位: {sum(1 for u in units if has_skill(u['talent_1']) or has_skill(u['talent_2']))}/{total}")
print(f"    有战斗技能 (battle) 的单位: {sum(1 for u in units if has_skill(u['battle_1']) or has_skill(u['battle_2']))}/{total}")
print(f"    仅有运营技能的单位:     {only_talent}/{total}")
print(f"    仅有战斗技能的单位:     {only_battle}/{total}")
print(f"    双阶段都有技能的单位:   {both_phases}/{total}")
print(f"    运营技能金色升级率:     {talent2_count}/{talent1_count}（有普通就有金色）")
print(f"    战斗技能金色升级率:     {battle2_count}/{battle1_count}")
print()

# ──────────────────────────────────────────────
# 二、触发方式分类
# ──────────────────────────────────────────────
print("━" * 70)
print("二、触发方式分类（Trigger Taxonomy）")
print("━" * 70)
print()

def classify_triggers(units):
    """将所有技能按触发词分类"""
    categories = {
        "运营—入场触发 (On Enter)": [],
        "运营—离场/出售触发 (On Leave/Sell)": [],
        "运营—回合结束时 (End of Turn)": [],
        "运营—回合开始时 (Start of Turn)": [],
        "运营—获得数量时 (On Gain Count)": [],
        "运营—累计N次触发 (Cumulative)": [],
        "运营—吞噬触发 (On Devour)": [],
        "运营—赐予触发 (On Bestow)": [],
        "运营—进阶触发 (On Evolve)": [],
        "运营—商店/手牌操控 (Shop/Hand Manip)": [],
        "战斗—开战后 (On Battle Start)": [],
        "战斗—每次攻击/攻击时 (On Attack)": [],
        "战斗—受到伤害/受到攻击 (On Damaged)": [],
        "战斗—阵亡时 (On Death)": [],
        "战斗—行动时 (On Action)": [],
        "战斗—造成追击后 (On Pursuit)": [],
        "战斗—受到攻击后存活 (After Attacked Survive)": [],
    }

    for u in units:
        unit = u["unit"]
        for col, text in [("talent_1", u["talent_1"]), ("talent_2", u["talent_2"]),
                          ("battle_1", u["battle_1"]), ("battle_2", u["battle_2"])]:
            if not has_skill(text):
                continue

            # 运营触发
            if "入场" in text and "吞噬" not in text and "商店" not in text:
                categories["运营—入场触发 (On Enter)"].append((unit, col, text))
            if "离场" in text or "出售时" in text:
                categories["运营—离场/出售触发 (On Leave/Sell)"].append((unit, col, text))
            if "回合结束时" in text:
                categories["运营—回合结束时 (End of Turn)"].append((unit, col, text))
            if "回合开始时" in text:
                categories["运营—回合开始时 (Start of Turn)"].append((unit, col, text))
            if ("获得数量" in text and "每获得" not in text) or "获得数量时" in text:
                categories["运营—获得数量时 (On Gain Count)"].append((unit, col, text))
            if "每获得" in text:
                categories["运营—获得数量时 (On Gain Count)"].append((unit, col, text))
            if "累计" in text:
                categories["运营—累计N次触发 (Cumulative)"].append((unit, col, text))
            if "吞噬" in text:
                categories["运营—吞噬触发 (On Devour)"].append((unit, col, text))
            if "赐予" in text:
                categories["运营—赐予触发 (On Bestow)"].append((unit, col, text))
            if "进阶" in text:
                categories["运营—进阶触发 (On Evolve)"].append((unit, col, text))
            if "商店" in text or "手牌" in text or "卡牌" in text:
                if "入场时，使商店" in text or "吞噬商店" in text or "获得" in text:
                    categories["运营—商店/手牌操控 (Shop/Hand Manip)"].append((unit, col, text))

            # 战斗触发
            if "开战后" in text:
                categories["战斗—开战后 (On Battle Start)"].append((unit, col, text))
            if "每次攻击" in text or "攻击时" in text:
                categories["战斗—每次攻击/攻击时 (On Attack)"].append((unit, col, text))
            if "受到攻击" in text or "受到伤害" in text:
                categories["战斗—受到伤害/受到攻击 (On Damaged)"].append((unit, col, text))
            if "阵亡" in text and "回合开始" not in text:
                categories["战斗—阵亡时 (On Death)"].append((unit, col, text))
            if "行动时" in text:
                categories["战斗—行动时 (On Action)"].append((unit, col, text))
            if "造成追击" in text:
                categories["战斗—造成追击后 (On Pursuit)"].append((unit, col, text))
            if "受到攻击后如果存活" in text or "受到攻击后" in text:
                categories["战斗—受到攻击后存活 (After Attacked Survive)"].append((unit, col, text))

    return categories


cats = classify_triggers(units)

for cat_name, items in cats.items():
    if not items:
        continue
    print(f"  ▸ {cat_name}  ({len(items)} 条技能)")
    for unit, col, txt in items[:5]:
        short = txt[:70] + ("..." if len(txt) > 70 else "")
        print(f"      · {unit} [{col}]: {short}")
    if len(items) > 5:
        print(f"      ... 还有 {len(items)-5} 条")
    print()

# ──────────────────────────────────────────────
# 三、效果类型分类
# ──────────────────────────────────────────────
print("━" * 70)
print("三、效果类型分类（Effect Type Taxonomy）")
print("━" * 70)
print()

def classify_effects(units):
    effects = {
        "数量增减 (Count +/-)": [],
        "百分比增减 (Percentage)": [],
        "攻击力增减 (ATK +/-)": [],
        "先机增减 (Initiative +/-)": [],
        "获得金币 (Gold)": [],
        "召唤/生成部队 (Summon)": [],
        "护盾 (Shield)": [],
        "暴击 (Critical)": [],
        "AOE范围伤害 (Area Damage)": [],
        "位移/冲锋 (Charge/Move)": [],
        "眩晕/控制 (Stun/CC)": [],
        "潜行 (Stealth)": [],
        "反击 (Counter)": [],
        "连续攻击 (Multi-attack)": [],
        "获得数量 (Gain Count)": [],
        "吞噬商店 (Devour Shop)": [],
        "进阶/变身 (Evolve/Transform)": [],
        "商店操控 (Shop Manip)": [],
        "手牌获取 (Card Draw)": [],
        "死亡后效果 (On-Death)": [],
        "多目标攻击 (Multi-target)": [],
    }

    for u in units:
        unit = u["unit"]
        for col, text in [("talent_1", u["talent_1"]), ("talent_2", u["talent_2"]),
                          ("battle_1", u["battle_1"]), ("battle_2", u["battle_2"])]:
            if not has_skill(text):
                continue
            if "数量" in text and ("+1" in text or "+2" in text or "+3" in text or "+4" in text or "+5" in text
                                   or "+10" in text or "+15" in text or "+20" in text or "获得" in text
                                   or "损失" in text or "少于" in text):
                effects["数量增减 (Count +/-)"].append((unit, col, text))
            if "%" in text or "百分比" in text:
                effects["百分比增减 (Percentage)"].append((unit, col, text))
            if "攻击" in text and ("临时攻击" in text or "攻击+" in text):
                effects["攻击力增减 (ATK +/-)"].append((unit, col, text))
            if "先机" in text:
                effects["先机增减 (Initiative +/-)"].append((unit, col, text))
            if "金币" in text:
                effects["获得金币 (Gold)"].append((unit, col, text))
            if "召唤" in text:
                effects["召唤/生成部队 (Summon)"].append((unit, col, text))
            if "护盾" in text:
                effects["护盾 (Shield)"].append((unit, col, text))
            if "暴击" in text:
                effects["暴击 (Critical)"].append((unit, col, text))
            if "范围" in text and ("伤害" in text or "攻击" in text):
                effects["AOE范围伤害 (Area Damage)"].append((unit, col, text))
            if ("冲到" in text or "猛扑" in text or "移动" in text or "锁住" in text) and "幻影" not in text:
                effects["位移/冲锋 (Charge/Move)"].append((unit, col, text))
            if "眩晕" in text or "锁住" in text:
                effects["眩晕/控制 (Stun/CC)"].append((unit, col, text))
            if "潜行" in text:
                effects["潜行 (Stealth)"].append((unit, col, text))
            if "反击" in text:
                effects["反击 (Counter)"].append((unit, col, text))
            if "连续攻击" in text:
                effects["连续攻击 (Multi-attack)"].append((unit, col, text))
            if "获得数量" in text:
                effects["获得数量 (Gain Count)"].append((unit, col, text))
            if "吞噬" in text and "商店" in text:
                effects["吞噬商店 (Devour Shop)"].append((unit, col, text))
            if "进阶" in text or "变为" in text or "变成" in text:
                effects["进阶/变身 (Evolve/Transform)"].append((unit, col, text))
            if "商店" in text and "卡牌" in text and "默认数量" in text:
                effects["商店操控 (Shop Manip)"].append((unit, col, text))
            if "获得" in text and "卡牌" in text:
                effects["手牌获取 (Card Draw)"].append((unit, col, text))
            if "阵亡后" in text or ("阵亡" in text and ("召唤" in text or "获得" in text or "产生" in text)):
                effects["死亡后效果 (On-Death)"].append((unit, col, text))
            if "额外" in text and ("支部队" in text or "敌人" in text) and "伤害" in text:
                effects["多目标攻击 (Multi-target)"].append((unit, col, text))

    return effects


effs = classify_effects(units)

for eff_name, items in effs.items():
    if not items:
        continue
    # 去重
    seen = set()
    unique = []
    for unit, col, txt in items:
        key = (unit, col)
        if key not in seen:
            seen.add(key)
            unique.append((unit, col, txt))
    print(f"  ▸ {eff_name}  ({len(unique)} 个技能)")
    for unit, col, txt in unique[:4]:
        short = txt[:70] + ("..." if len(txt) > 70 else "")
        print(f"      · {unit} [{col}]: {short}")
    if len(unique) > 4:
        print(f"      ... 还有 {len(unique)-4} 个")
    print()

# ──────────────────────────────────────────────
# 四、种族/信仰体系与技能关联
# ──────────────────────────────────────────────
print("━" * 70)
print("四、种族/信仰体系与技能关联（Faction Mechanic Mapping）")
print("━" * 70)
print()

# 统计各faith/race
faith_groups = defaultdict(list)
for u in units:
    faith_groups[u["faith"]].append(u)

for faith, members in faith_groups.items():
    if faith == "faith":
        continue
    skill_units = [m for m in members if has_skill(m["talent_1"]) or has_skill(m["talent_2"])
                   or has_skill(m["battle_1"]) or has_skill(m["battle_2"])]
    print(f"  ▸ 信仰: {faith}  ({len(members)} 单位, {len(skill_units)} 有技能)")
    # 提取该信仰的独有机制关键词
    all_text = " ".join([m["talent_1"] + " " + m["talent_2"] + " " + m["battle_1"] + " " + m["battle_2"] for m in members])
    keywords = []
    if "莱特" in all_text:
        keywords.append("莱特信仰者协同")
    if "甘地" in all_text:
        keywords.append("甘地部队计数触发")
    if "甘德" in all_text:
        keywords.append("甘德部队关联")
    if "甘格尔" in all_text:
        keywords.append("甘格尔吞噬机制")
    if "元素" in all_text:
        keywords.append("元素种族链")
    if "密林宝钻" in all_text:
        keywords.append("密林宝钻赐予系统")
    if "野兽" in all_text:
        keywords.append("野兽入场链")
    if "魔灵" in all_text:
        keywords.append("魔灵离场链")
    if "艾瑞" in all_text:
        keywords.append("艾瑞信仰关联")
    if "吞噬" in all_text:
        keywords.append("吞噬商店")
    if "法师" in all_text:
        keywords.append("法师标签")
    if "战士" in all_text:
        keywords.append("战士标签")
    if "获得数量" in all_text:
        keywords.append("获得数量机制")
    if "入场" in all_text:
        keywords.append("入场效果")
    if keywords:
        print(f"     核心机制: {', '.join(keywords)}")
    print()

# ──────────────────────────────────────────────
# 五、条件判定体系
# ──────────────────────────────────────────────
print("━" * 70)
print("五、条件判定体系（Conditional System）")
print("━" * 70)
print()

conditions = {
    "场上数量条件 (Field Count ≥ N)": [],
    "相邻/同排/前后排 (Position Adjacency)": [],
    "当前数量阈值 (Self Count Threshold)": [],
    "信仰/种族标签过滤 (Faith/Race Filter)": [],
    "首次攻击限定 (First Attack Only)": [],
    "每回合限N次 (Per Turn Limit)": [],
    "每场战斗限N次 (Per Battle Limit)": [],
    "自身存活判定 (Survive Check)": [],
    "距离条件 (Distance Check)": [],
    "数量超过初始N倍 (Count Multiplier)": [],
}

for u in units:
    unit = u["unit"]
    for col, text in [("talent_1", u["talent_1"]), ("talent_2", u["talent_2"]),
                      ("battle_1", u["battle_1"]), ("battle_2", u["battle_2"])]:
        if not has_skill(text):
            continue
        if "场上至少" in text or "场上.*达" in text or "每有1" in text or "每有1支" in text:
            conditions["场上数量条件 (Field Count ≥ N)"].append((unit, col, text))
        if "相邻" in text or "同一排" in text or "前一排" in text or "后一排" in text or "左右两侧" in text or "最近" in text:
            conditions["相邻/同排/前后排 (Position Adjacency)"].append((unit, col, text))
        if ("数量≥" in text or "数量≥" in text or "数量少于" in text or "数量超过" in text
                or "数量不足" in text):
            conditions["当前数量阈值 (Self Count Threshold)"].append((unit, col, text))
        if "莱特" in text or "甘地" in text or "甘德" in text or "甘格尔" in text or "元素" in text or "野兽" in text or "魔灵" in text or "艾瑞" in text or "法师" in text or "战士" in text:
            conditions["信仰/种族标签过滤 (Faith/Race Filter)"].append((unit, col, text))
        if "首次攻击" in text or "第1次" in text or "前2次" in text:
            conditions["首次攻击限定 (First Attack Only)"].append((unit, col, text))
        if "每回合" in text and "限" in text:
            conditions["每回合限N次 (Per Turn Limit)"].append((unit, col, text))
        if "每场战斗" in text:
            conditions["每场战斗限N次 (Per Battle Limit)"].append((unit, col, text))
        if "如果存活" in text:
            conditions["自身存活判定 (Survive Check)"].append((unit, col, text))
        if "距离不少于" in text or "距离" in text:
            conditions["距离条件 (Distance Check)"].append((unit, col, text))
        if "2倍" in text or "3倍" in text or "6倍" in text:
            if "数量" in text:
                conditions["数量超过初始N倍 (Count Multiplier)"].append((unit, col, text))

for cond_name, items in conditions.items():
    if not items:
        continue
    seen = set()
    unique = []
    for unit, col, txt in items:
        key = (unit, col)
        if key not in seen:
            seen.add(key)
            unique.append((unit, col, txt))
    print(f"  ▸ {cond_name}  ({len(unique)} 个技能)")
    for unit, col, txt in unique[:3]:
        short = txt[:70] + ("..." if len(txt) > 70 else "")
        print(f"      · {unit} [{col}]: {short}")
    if len(unique) > 3:
        print(f"      ... 还有 {len(unique)-3} 个")
    print()

# ──────────────────────────────────────────────
# 六、数值增长模型
# ──────────────────────────────────────────────
print("━" * 70)
print("六、数值增长模型（Scaling Model）")
print("━" * 70)
print()
print("  金色升级(talent_2 / battle_2)相对普通形态(talent_1 / battle_1)的数值增幅分析：")
print()
print("  增幅模式归纳：")
print("    · 直接翻倍型：普通+1→金色+2, +2→+4, +4→+8, +5→+10 (最常见)")
print("    · 触发次数翻倍型：触发1次→触发2次, 1支部队→2支部队")
print("    · 百分比提升型：10%→20%, 20%→30%, 25%→35%概率")
print("    · 范围扩大型：周围1距离→2距离, 2格→3格→4格")
print("    · 阈值降低型：累计10次→累计5次 (触发更容易)")
print("    · 倍率提高型：3倍暴击→6倍暴击")
print("    · 护盾层数型：抵挡1次→抵挡2次")
print("    · 持续时间型：持续2轮→持续3轮")
print("    · 属性提升型：临时+2先机→临时+4先机, +1攻击→+2攻击")
print()
print("  关键洞察：金色升级的数值设计以 '2倍线性增幅' 为主基调，")
print("  少数技能采用 '范围扩大' 或 '条件放宽' 的非线性增强方式。")
print()

# ──────────────────────────────────────────────
# 七、独特机制词条
# ──────────────────────────────────────────────
print("━" * 70)
print("七、独特机制词条（Unique Mechanic Keywords）")
print("━" * 70)
print()
unique_mechanics = {
    "密林宝钻 (Emerald Gem)": "独立资源系统，通过赐予/获得/消费形成闭环经济，驱动整个林地种族",
    "吞噬商店 (Devour Shop)": "格尔种族独有，消耗商店卡牌转化为数量，形成独特的资源转化链",
    "入场效果重触发 (Re-trigger Battlecry)": "鱼人奴仆、劣徒可让其他部队重新触发入场效果，形成Combo引擎",
    "进阶系统 (Evolution)": "满足条件后单位变为更强形态（游骑兵→精锐游骑兵，游侠→神剑游侠等）",
    "潜行 (Stealth)": "战斗内首次攻击必定暴击，改变战斗节奏",
    "护盾 (Shield)": "完全抵挡伤害的防护机制，不可叠加",
    "幻影部队 (Phantom)": "召唤不可移动/攻击的肉盾，改变战场布局",
    "冲锋 (Charge)": "战斗开始瞬间位移到敌人面前，打破阵型",
    "骑乘变身 (Mount Transform)": "邪恶女巫骑上格尔兽变为巫兽师，战斗内形态切换",
    "临时数量 (Temp Count)": "战斗中临时获得的增益，战后不保留",
    "获得数量 (Gain Count)": "触发友军数量增长的机制",
    "连续攻击 (Multi-Attack)": "一回合内攻击多次，极大提升DPS",
    "火雨AOE (Rain of Fire)": "以敌人位置为中心的范围伤害",
    "爆炸亡语 (Deathrattle Explosion)": "阵亡时产生范围爆炸",
}
for mech, desc in unique_mechanics.items():
    print(f"  ▸ {mech}")
    print(f"      {desc}")
    print()

# ──────────────────────────────────────────────
# 八、设计总结
# ──────────────────────────────────────────────
print("━" * 70)
print("八、底层设计哲学总结")
print("━" * 70)
print()
print("  1. 【严格的阶段分离】")
print("     运营技能只影响局外经济/成长，战斗技能只影响局内战局。")
print("     互不交叉，确保玩家在\"构筑阶段\"和\"战斗阶段\"有清晰的决策边界。")
print()
print("  2. 【种族即机制 (Race = Mechanic)】")
print("     每个信仰/种族定义了一套独特的资源循环机制：")
print("       · 甘地 → 部队数量计数触发链")
print("       · 莱特 → 信仰者协同获得数量")
print("       · 元素 → 种族内数量/离场链")
print("       · 林地 → 密林宝钻赐予经济系统")
print("       · 格尔 → 吞噬商店资源转化")
print("     这使得每个种族有完全不同的玩法风格。")
print()
print("  3. 【金色升级 = 2倍线性增幅】")
print("     普通→金色的数值设计简单清晰：大部分是 x2 翻倍。")
print("     少数高级单位有范围扩大/条件放宽的质变。")
print("     这降低了玩家理解成本，同时保持了升级的满足感。")
print()
print("  4. 【触发链与Combo设计】")
print("     技能之间存在明显的触发链设计：")
print("       · 入场→入场重触发→获得数量→累计触发→进阶")
print("       · 吞噬→吞噬事件→格尔巨兽/邪恶女巫受益")
print("       · 赐予密林宝钻→赐予累计→林地将军/猎豹触发")
print("     这些链式反应是构筑深度的核心来源。")
print()
print("  5. 【条件判定的渐进复杂度】")
print("     低级单位（星级1-2）：简单条件（回合结束、入场、离场）")
print("     中级单位（星级3-4）：中等条件（累计N次、场上数量阈值、位置关系）")
print("     高级单位（星级5-6）：复杂条件（多条件组合、跨阶段效果、形态切换）")
print("     这种渐进设计符合玩家学习曲线。")
print()
print("  6. 【战斗技能的战术维度】")
print("     战斗技能覆盖了完整的战术维度：")
print("       · 先手爆发：开战后召唤/冲锋/暴击")
print("       · 持续输出：连续攻击/AOE/多目标")
print("       · 防御反制：护盾/反击/潜行")
print("       · 控场与位移：眩晕/锁住/冲锋")
print("       · 阵亡补偿：亡语爆炸/亡语召唤/亡语资源")
print("       · 成长型：行动时累计增益/受击后反击")
print()
print("=" * 70)
print("                    分析完毕")
print("=" * 70)
