# -*- coding: utf-8 -*-
"""Restore original (unboosted) values for 5 non-boss presets. Keep Boss P0 fix."""
import openpyxl

EXCEL_UNITS = "docs/excel/怪物配置_敌方单位明细表.xlsx"

RESTORE = {
    "shadow_raiders": [
        ("enemy_1", "assassin", 14, 3, "前排爆发", "刺客暴击 检验后排承伤"),
        ("enemy_2", "wanderer", 12, 3, "副输出", "流浪者 中距输出"),
        ("enemy_3", "blacksmith", 14, 2, "辅助输出", "铁匠 补充近战"),
        ("enemy_4", "light_mentor", 6, 4, "后排召唤", "光明导师 召唤幻影"),
    ],
    "garrison_line": [
        ("enemy_1", "garrison_guard", 5, 5, "前排核心", "卫戍协兵 全军增益联动"),
        ("enemy_2", "priest", 9, 3, "后排治疗", "牧师 护盾支持"),
        ("enemy_3", "light_envoy", 5, 5, "后排支援", "莱特使者 信仰协同"),
        ("enemy_4", "wanderer", 12, 3, "副输出", "流浪者 中距持续输出"),
    ],
    "abyss_vanguard": [
        ("enemy_1", "martial_master", 8, 4, "前排AOE", "武学大师 AOE范围伤害"),
        ("enemy_2", "light_mentor", 7, 4, "后排召唤", "光明导师 召唤幻影部队"),
        ("enemy_3", "assassin", 14, 3, "刺客爆发", "刺客 潜行暴击"),
        ("enemy_4", "garrison_guard", 5, 5, "辅助增益", "卫戍协兵 提供全军buff"),
    ],
    "fallen_sanctum": [
        ("enemy_1", "light_envoy", 5, 5, "核心支援", "莱特使者 信仰协同核心"),
        ("enemy_2", "garrison_guard", 6, 5, "前排承伤", "卫戍协兵 高防联动"),
        ("enemy_3", "light_mentor", 8, 4, "后排召唤", "光明导师 多单位召唤"),
        ("enemy_4", "royal_swordsman", 3, 6, "AOE输出", "皇家剑士 战士AOE"),
    ],
    "doom_herald": [
        ("enemy_1", "royal_swordsman", 6, 6, "前排核心", "皇家剑士 战士增益"),
        ("enemy_2", "echo_of_light", 3, 6, "后排输出", "莱特回响 开战全军增益"),
        ("enemy_3", "martial_master", 8, 5, "前排AOE", "武学大师 数量压制"),
        ("enemy_4", "garrison_guard", 6, 5, "辅助增益", "卫戍协兵 终局协同检查"),
    ],
}

wb = openpyxl.load_workbook(EXCEL_UNITS)
ws = wb.active

preset_rows = {}
for row in range(2, ws.max_row + 1):
    pid = str(ws.cell(row=row, column=1).value or "").strip()
    if pid in RESTORE:
        if pid not in preset_rows:
            preset_rows[pid] = []
        preset_rows[pid].append(row)

for pid, slots in RESTORE.items():
    rows = preset_rows.get(pid, [])
    for i, row_num in enumerate(sorted(rows)):
        slot_id, unit_id, count, star, role, notes = slots[i]
        ws.cell(row=row_num, column=2).value = slot_id
        ws.cell(row=row_num, column=3).value = unit_id
        ws.cell(row=row_num, column=4).value = count
        ws.cell(row=row_num, column=5).value = star
        ws.cell(row=row_num, column=6).value = role
        ws.cell(row=row_num, column=7).value = notes
    print(f"✅ Restored {pid}")

wb.save(EXCEL_UNITS)
print(f"\nSaved {EXCEL_UNITS}")
print("\n⚠ Boss abyss_lord KEPT at boosted 1486 (P0 fix)")
print("  5 non-boss presets RESTORED to original values")