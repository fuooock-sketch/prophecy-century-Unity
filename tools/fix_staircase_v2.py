# -*- coding: utf-8 -*-
"""Fix enemy staircase V2 - moderate boosts."""
import openpyxl

EXCEL_UNITS = "docs/excel/怪物配置_敌方单位明细表.xlsx"

V2 = {
    "shadow_raiders": [
        ("enemy_1", "assassin", 16, 3, "前排爆发", "刺客暴击 检验后排承伤"),
        ("enemy_2", "wanderer", 13, 3, "副输出", "流浪者 中距输出"),
        ("enemy_3", "blacksmith", 15, 3, "辅助输出", "铁匠 ★3填充"),
        ("enemy_4", "light_mentor", 7, 4, "后排召唤", "光明导师 召唤幻影"),
    ],
    "garrison_line": [
        ("enemy_1", "garrison_guard", 6, 5, "前排核心", "卫戍协兵 全军增益联动"),
        ("enemy_2", "priest", 11, 3, "后排治疗", "牧师 护盾支持"),
        ("enemy_3", "light_envoy", 6, 5, "后排支援", "莱特使者 信仰协同"),
        ("enemy_4", "wanderer", 14, 3, "副输出", "流浪者 中距输出"),
    ],
    "abyss_vanguard": [
        ("enemy_1", "martial_master", 9, 5, "前排AOE", "武学大师 ★5范围伤害"),
        ("enemy_2", "light_mentor", 8, 5, "后排召唤", "光明导师 ★5召唤幻影"),
        ("enemy_3", "assassin", 16, 4, "刺客爆发", "刺客 ★4潜行暴击"),
        ("enemy_4", "garrison_guard", 6, 6, "辅助增益", "卫戍协兵 ★6全军buff"),
    ],
    "fallen_sanctum": [
        ("enemy_1", "light_envoy", 6, 5, "核心支援", "莱特使者 信仰协同"),
        ("enemy_2", "garrison_guard", 7, 6, "前排承伤", "卫戍协兵 ★6高防联动"),
        ("enemy_3", "light_mentor", 9, 5, "后排召唤", "光明导师 ★5多单位召唤"),
        ("enemy_4", "royal_swordsman", 4, 6, "AOE输出", "皇家剑士 ★6战士AOE"),
    ],
    "doom_herald": [
        ("enemy_1", "royal_swordsman", 7, 6, "前排核心", "皇家剑士 ★6战士增益"),
        ("enemy_2", "echo_of_light", 4, 6, "后排输出", "莱特回响 开战全军增益"),
        ("enemy_3", "martial_master", 9, 5, "前排AOE", "武学大师 ★5数量压制"),
        ("enemy_4", "garrison_guard", 7, 6, "辅助增益", "卫戍协兵 ★6终局协同"),
    ],
}

wb = openpyxl.load_workbook(EXCEL_UNITS)
ws = wb.active

# Find rows
preset_rows = {}
for row in range(2, ws.max_row + 1):
    pid = str(ws.cell(row=row, column=1).value or "").strip()
    if pid in V2:
        if pid not in preset_rows:
            preset_rows[pid] = []
        preset_rows[pid].append(row)

for pid, new_slots in V2.items():
    rows = preset_rows.get(pid, [])
    if len(rows) != len(new_slots):
        print(f"⚠ {pid}: expected {len(new_slots)} rows, found {len(rows)}")
        continue
    for i, row_num in enumerate(sorted(rows)):
        slot_id, unit_id, count, star, role, notes = new_slots[i]
        ws.cell(row=row_num, column=2).value = slot_id
        ws.cell(row=row_num, column=3).value = unit_id
        ws.cell(row=row_num, column=4).value = count
        ws.cell(row=row_num, column=5).value = star
        ws.cell(row=row_num, column=6).value = role
        ws.cell(row=row_num, column=7).value = notes
    print(f"✅ {pid} ({len(rows)} slots)")

wb.save(EXCEL_UNITS)
print(f"Saved {EXCEL_UNITS}")
print("Done!")